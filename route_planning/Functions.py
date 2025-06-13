#!/usr/bin/env python
# coding=utf-8
# This file is used to further process data and visualize it.


import json
import math
import pandas as pd
import numpy as np
from xlwt import *
import openpyxl
import matplotlib.pyplot as plt
from scipy.interpolate import make_interp_spline


# input dict
# output boxplot
def dict_to_boxplot(dict1, dict2, dict3, dict4, title):
    data = [list(dict1.values()), list(dict2.values()), list(dict3.values()), list(dict4.values())]

    plt.figure(figsize=(10, 6))
    plt.boxplot(data, labels=['Baseline', 'QR', 'CQR', 'CQR-ERC'])
    plt.title(title)
    plt.ylabel('Values')
    plt.show()
    return 0


# input dict
# output scatterplot
def dict_scatter(dict1, dict2, dict3, dict4):
    # 四个输入的列表数据
    list1 = list(dict1.values())
    list2 = list(dict2.values())
    list3 = list(dict3.values())
    list4 = list(dict4.values())

    # 创建一个新的图形
    plt.figure(figsize=(10, 6))

    # 绘制散点图
    plt.scatter([1] * len(list1), list1, label='Baseline')
    plt.scatter([2] * len(list2), list2, label='QR')
    plt.scatter([3] * len(list3), list3, label='CQR')
    plt.scatter([4] * len(list4), list4, label='CQR-ERC')

    # 添加图例
    plt.legend()

    plt.title('Scatter Plot of 4 types')
    plt.xlabel('List Number')
    plt.ylabel('Values')
    plt.xticks([1, 2, 3, 4], ['1', '2', '3', '4'])

    plt.show()
    return 0


# input 7 num_lists
# output plot
def plot_label_and_interval(numlist5, numlist6, numlist7):
    # 创建绘图
    y = list(range(len(numlist7)))

    plt.figure(figsize=(10, 6))

    # 绘制散点图

    plt.scatter(y, numlist5, color='green', marker='^', label='qr', s=3)
    plt.scatter(y, numlist6, color='green', marker='^', s=3)

    plt.scatter(y, numlist7, color='red', marker='D', label='edge_weight', s=5)

    # 设置图例

    plt.legend()

    # 设置标题和标签

    plt.title('Scatter_plot')

    plt.xlabel('X_label')

    plt.ylabel('Y_label')

    # 显示图形

    plt.show()

    return 0


def del_p_key_in_file(file1, file2, file3, file4, file5, file6, file7, file8, file9, file10, file11, file12, file13,
                      file14, key, NNNN):
    obj_cqr = json.load(open(file1))
    obj_qr = json.load(open(file2))
    obj_bl = json.load(open(file3))
    obj_cqr_label = json.load(open(file4))
    obj_qr_label = json.load(open(file5))
    obj_bl_label = json.load(open(file6))
    obj_cqr_new = json.load(open(file7))
    obj_cqr_new_label = json.load(open(file8))
    cov_cqr = json.load(open(file9))
    cov_cqr_new = json.load(open(file10))
    cov_qr = json.load(open(file11))
    x_cqr = json.load(open(file12))
    x_cqr_new = json.load(open(file13))
    x_qr = json.load(open(file14))

    for k in [obj_cqr, obj_qr, obj_bl, obj_cqr_label, obj_qr_label, obj_bl_label, obj_cqr_new, obj_cqr_new_label,
              cov_cqr, cov_cqr_new, cov_qr, x_cqr, x_cqr_new, x_qr]:
        for key1, value1 in k.items():
            for key2, value2 in value1.items():
                if key in value2:
                    del value1[key2][key]

    dump_files(NNNN, obj_cqr, obj_cqr_new, obj_qr, obj_bl, obj_cqr_label, obj_cqr_new_label, obj_qr_label, obj_bl_label,
               cov_cqr, cov_cqr_new, cov_qr, x_cqr, x_cqr_new, x_qr)
    return 0


def del_seed_key_in_file(file1, file2, file3, file4, file5, file6, file7, file8, file9, file10, file11, file12, file13,
                         file14, key, NNNN):
    obj_cqr = json.load(open(file1))
    obj_qr = json.load(open(file2))
    obj_bl = json.load(open(file3))
    obj_cqr_label = json.load(open(file4))
    obj_qr_label = json.load(open(file5))
    obj_bl_label = json.load(open(file6))
    obj_cqr_new = json.load(open(file7))
    obj_cqr_new_label = json.load(open(file8))
    cov_cqr = json.load(open(file9))
    cov_cqr_new = json.load(open(file10))
    cov_qr = json.load(open(file11))
    x_cqr = json.load(open(file12))
    x_cqr_new = json.load(open(file13))
    x_qr = json.load(open(file14))

    for k in [obj_cqr, obj_qr, obj_bl, obj_cqr_label, obj_qr_label, obj_bl_label, obj_cqr_new, obj_cqr_new_label,
              cov_cqr, cov_cqr_new, cov_qr, x_cqr, x_cqr_new, x_qr]:
        for keys, value in k.items():
            if key in value:
                del value[key]

    dump_files(NNNN, obj_cqr, obj_cqr_new, obj_qr, obj_bl, obj_cqr_label, obj_cqr_new_label, obj_qr_label, obj_bl_label,
               cov_cqr, cov_cqr_new, cov_qr, x_cqr, x_cqr_new, x_qr)
    return 0


def del_seed_key_in_file(file1, file2, file3, file4, file5, file6, file7, file8, file9, key, NEW_NAME):
    CQR_interval = json.load(open(file1))
    CQR_NEW_interval = json.load(open(file2))
    QR_interval = json.load(open(file3))
    BL_interval = json.load(open(file4))
    C_LABEL = json.load(open(file5))
    COV_CQR = json.load(open(file6))
    COV_CQR_NEW = json.load(open(file7))
    COV_QR = json.load(open(file8))
    COV_BL = json.load(open(file9))

    for k in [CQR_interval, CQR_NEW_interval, QR_interval, BL_interval, C_LABEL, COV_CQR, COV_CQR_NEW, COV_QR,
              COV_BL]:
        for keys, value in k.items():
            if key in value:
                del value[key]

    dump_files_NO_OPT(NEW_NAME, CQR_interval, CQR_NEW_interval, QR_interval, BL_interval, C_LABEL, COV_CQR, COV_CQR_NEW,
                      COV_QR, COV_BL)

    return 0


def del_ST_key_in_file(file_obj_cqr, file_obj_cqr_new, file_obj_qr, file_obj_bl, file_cqr_label, file_cqr_new_label,
                       file_qr_label, file_bl_label, file_cov_cqr_new, file_cov_cqr, file_cov_qr, file_cov_bl,
                       file_x_cqr, file_x_cqr_new, file_x_bl, file_x_qr, key, NNNN):
    obj_cqr = json.load(open(file_obj_cqr))
    obj_qr = json.load(open(file_obj_qr))
    obj_bl = json.load(open(file_obj_bl))
    obj_cqr_label = json.load(open(file_cqr_label))
    obj_qr_label = json.load(open(file_qr_label))
    obj_bl_label = json.load(open(file_bl_label))
    obj_cqr_new = json.load(open(file_obj_cqr_new))
    obj_cqr_new_label = json.load(open(file_cqr_new_label))
    cov_cqr = json.load(open(file_cov_cqr))
    cov_cqr_new = json.load(open(file_cov_cqr_new))
    cov_qr = json.load(open(file_cov_qr))
    cov_bl = json.load(open(file_cov_bl))
    x_cqr = json.load(open(file_x_cqr))
    x_cqr_new = json.load(open(file_x_cqr_new))
    x_qr = json.load(open(file_x_qr))
    x_bl = json.load(open(file_x_bl))

    for k in [obj_cqr, obj_qr, obj_bl, obj_cqr_label, obj_qr_label, obj_bl_label, obj_cqr_new, obj_cqr_new_label,
              cov_cqr, cov_cqr_new, cov_qr, cov_bl, x_cqr, x_cqr_new, x_qr, x_bl]:
        if key in k:
            del k[key]

    dump_files(NNNN, obj_cqr, obj_cqr_new, obj_qr, obj_bl, obj_cqr_label, obj_cqr_new_label, obj_qr_label, obj_bl_label,
               cov_cqr, cov_cqr_new, cov_qr, cov_bl, x_cqr, x_cqr_new, x_qr, x_bl)
    return 0


def dump_files(NAME, OBJ_CQR, OBJ_CQR_NEW, OBJ_QR, OBJ_BL,
               OBJ_CQR_LABEL, OBJ_CQR_NEW_LABEL, OBJ_QR_LABEL,
               OBJ_BL_LABEL, COV_CQR, COV_CQR_NEW, COV_QR, COV_BL,
               X_CQR, X_CQR_NEW, X_QR, X_BL):
    json.dump(X_CQR, open(
        r'D:\OneDrive - City University of Hong Kong - Student\PyCharm202122\shortest_path202403\GNN\Data' + NAME + '\X_CQR.json',
        'w'))
    json.dump(X_CQR_NEW, open(
        r'D:\OneDrive - City University of Hong Kong - Student\PyCharm202122\shortest_path202403\GNN\Data' + NAME + '\X_CQR_NEW.json',
        'w'))
    json.dump(X_QR, open(
        r'D:\OneDrive - City University of Hong Kong - Student\PyCharm202122\shortest_path202403\GNN\Data' + NAME + '\X_QR.json',
        'w'))
    json.dump(X_BL, open(
        r'D:\OneDrive - City University of Hong Kong - Student\PyCharm202122\shortest_path202403\GNN\Data' + NAME + '\X_BL.json',
        'w'))
    json.dump(OBJ_CQR, open(
        r'D:\OneDrive - City University of Hong Kong - Student\PyCharm202122\shortest_path202403\GNN\Data' + NAME + '\OBJ_CQR.json',
        'w'))
    json.dump(OBJ_CQR_NEW, open(
        r'D:\OneDrive - City University of Hong Kong - Student\PyCharm202122\shortest_path202403\GNN\Data' + NAME + '\OBJ_CQR_NEW.json',
        'w'))
    json.dump(OBJ_QR, open(
        r'D:\OneDrive - City University of Hong Kong - Student\PyCharm202122\shortest_path202403\GNN\Data' + NAME + '\OBJ_QR.json',
        'w'))
    json.dump(OBJ_BL, open(
        r'D:\OneDrive - City University of Hong Kong - Student\PyCharm202122\shortest_path202403\GNN\Data' + NAME + '\OBJ_BL.json',
        'w'))
    json.dump(OBJ_CQR_LABEL, open(
        r'D:\OneDrive - City University of Hong Kong - Student\PyCharm202122\shortest_path202403\GNN\Data' + NAME + '\OBJ_CQR_LABEL.json',
        'w'))
    json.dump(OBJ_CQR_NEW_LABEL, open(
        r'D:\OneDrive - City University of Hong Kong - Student\PyCharm202122\shortest_path202403\GNN\Data' + NAME + '\OBJ_CQR_NEW_LABEL.json',
        'w'))
    json.dump(OBJ_QR_LABEL, open(
        r'D:\OneDrive - City University of Hong Kong - Student\PyCharm202122\shortest_path202403\GNN\Data' + NAME + '\OBJ_QR_LABEL.json',
        'w'))
    json.dump(OBJ_BL_LABEL, open(
        r'D:\OneDrive - City University of Hong Kong - Student\PyCharm202122\shortest_path202403\GNN\Data' + NAME + '\OBJ_BL_LABEL.json',
        'w'))
    json.dump(COV_CQR, open(
        r'D:\OneDrive - City University of Hong Kong - Student\PyCharm202122\shortest_path202403\GNN\Data' + NAME + '\COV_CQR.json',
        'w'))
    json.dump(COV_CQR_NEW, open(
        r'D:\OneDrive - City University of Hong Kong - Student\PyCharm202122\shortest_path202403\GNN\Data' + NAME + '\COV_CQR_NEW.json',
        'w'))
    json.dump(COV_QR, open(
        r'D:\OneDrive - City University of Hong Kong - Student\PyCharm202122\shortest_path202403\GNN\Data' + NAME + '\COV_QR.json',
        'w'))
    json.dump(COV_BL, open(
        r'D:\OneDrive - City University of Hong Kong - Student\PyCharm202122\shortest_path202403\GNN\Data' + NAME + '\COV_BL.json',
        'w'))


def dump_files_NO_OPT(NAME, CQR_interval, CQR_NEW_interval, QR_interval, BL_interval, C_LABEL, COV_CQR, COV_CQR_NEW,
                      COV_QR, COV_BL):
    json.dump(CQR_interval, open(
        r'D:\OneDrive - City University of Hong Kong - Student\PyCharm202122\shortest_path202403\GNN\Data' + NAME + '\CQR_interval.json',
        'w'))
    json.dump(CQR_NEW_interval, open(
        r'D:\OneDrive - City University of Hong Kong - Student\PyCharm202122\shortest_path202403\GNN\Data' + NAME + '\CQR_NEW_interval.json',
        'w'))
    json.dump(QR_interval, open(
        r'D:\OneDrive - City University of Hong Kong - Student\PyCharm202122\shortest_path202403\GNN\Data' + NAME + '\QR_interval.json',
        'w'))
    json.dump(BL_interval, open(
        r'D:\OneDrive - City University of Hong Kong - Student\PyCharm202122\shortest_path202403\GNN\Data' + NAME + '\BL_interval.json',
        'w'))
    json.dump(C_LABEL, open(
        r'D:\OneDrive - City University of Hong Kong - Student\PyCharm202122\shortest_path202403\GNN\Data' + NAME + '\C_LABEL.json',
        'w'))
    json.dump(COV_CQR, open(
        r'D:\OneDrive - City University of Hong Kong - Student\PyCharm202122\shortest_path202403\GNN\Data' + NAME + '\COV_CQR.json',
        'w'))
    json.dump(COV_CQR_NEW, open(
        r'D:\OneDrive - City University of Hong Kong - Student\PyCharm202122\shortest_path202403\GNN\Data' + NAME + '\COV_CQR_NEW.json',
        'w'))
    json.dump(COV_QR, open(
        r'D:\OneDrive - City University of Hong Kong - Student\PyCharm202122\shortest_path202403\GNN\Data' + NAME + '\COV_QR.json',
        'w'))
    json.dump(COV_BL, open(
        r'D:\OneDrive - City University of Hong Kong - Student\PyCharm202122\shortest_path202403\GNN\Data' + NAME + '\COV_BL.json',
        'w'))


def cost_obj_dic_to_xlsx_sheet(dict, filename, sheetname):
    file = Workbook(encoding='utf-8')
    # 指定file以utf-8的格式打开
    table = file.add_sheet(sheetname)
    # 指定打开的文件名
    # 字典数据
    ldata = []
    num = [a for a in dict]
    # for循环指定取出key值存入num中
    # num.sort()
    # 字典数据取出后无需，需要先排序

    for x in num:
        # for循环将data字典中的键和值分批的保存在ldata中
        t = [int(x)]
        if sheetname == 'cost':
            for a in dict[x]:
                t.append(a)
            ldata.append(t)
        else:
            t.append(dict[x])
            ldata.append(t)
    for i, p in enumerate(ldata):
        # 将数据写入文件,i是enumerate()函数返回的序号数
        for j, q in enumerate(p):
            # print i,j,q
            table.write(i, j, q)
    file.save(filename)
    return 0


def x_dic_to_xlxs_sheet(dict, filename, sheetname):
    # 获取工作簿对象
    wb = openpyxl.load_workbook(filename)

    # 获取工作表对象
    sheet = wb[sheetname]
    # 修改指定单元格内容
    for i in dict.keys():
        for j in range(0, len(dict[i])):
            index = str(dict[i][j][0]) + ',' + str(dict[i][j][1])
            sheet.cell(row=int(i) + 1, column=j + 1, value=index)
    # 写入多个单元格(追加模式，不会覆盖之前的，从有数据的下一行开始)
    # sheet.append(['王五', '三年级二班', '10岁'])

    # 复制"学生表"，新sheet名称为"学生表 Copy"
    # ws_2 = wb.copy_worksheet(wb['学生表'])
    # 保存
    wb.save(filename)
    return 0


def calculate_standard_error(numbers, invalid):
    n = len(numbers)
    for i in range(0, invalid):
        numbers.remove(0.1)
    mean = sum(numbers) / n
    squared_diff_sum = sum((x - mean) ** 2 for x in numbers)
    standard_error = math.sqrt(squared_diff_sum / n)
    return standard_error


def json_dic_to_std_error(FILENAME):
    dic = json.load(open(FILENAME))
    invalid = 0
    num_list = []
    for i in dic.keys():
        for j in dic[i].keys():
            for k in dic[i][j].keys():
                for l in dic[i][j][k].keys():
                    num_list.append(dic[i][j][k][l])
                    if dic[i][j][k][l] == 0.1:
                        invalid += 1
    std_e = calculate_standard_error(num_list, invalid)
    mean = (sum(num_list) - 0.1 * invalid) / (len(num_list) - invalid)
    return std_e, mean


def check_range(lower_array, upper_array, label_array):
    count = 0
    total = len(label_array)

    for lower, upper, label in zip(lower_array, upper_array, label_array):
        if lower <= label <= upper:
            count += 1

    ratio = count / total
    return ratio


def json_dic_bl_interval_to_cov(BL_NAME, LABEL_NAME):
    bl_dic = json.load(open(BL_NAME))
    label_dic = json.load(open(LABEL_NAME))
    invalid = 0
    num_list = []
    for i in bl_dic.keys():
        for j in bl_dic[i].keys():
            for k in bl_dic[i][j].keys():
                for l in bl_dic[i][j][k].keys():
                    num_list.append(check_range(np.zeros(len(label_dic[i][j][k][l])), np.array(bl_dic[i][j][k][l][0]),
                                                np.array(label_dic[i][j][k][l])))

    mean = np.mean(num_list)
    std = np.std(num_list)

    return std, mean


# key: p
def json_shuffle_to_std_error_dic(FILENAME):
    dic = json.load(open(FILENAME))
    std_error_dic_10x10x10 = {}
    mean_10x10x10 = {}
    for i in dic.keys():
        std_error_dic_10x10 = {}
        mean_10x10 = {}
        for j in dic[i].keys():
            std_error_dic_10 = {}
            mean_10 = {}
            for k in dic[i][j].keys():
                invalid = 0
                num_list_p = []
                for l in dic[i][j][k].keys():
                    num_list_p.append(dic[i][j][k][l])
                    if dic[i][j][k][l] == 0.1:
                        invalid += 1
                std_e = calculate_standard_error(num_list_p, invalid)
                std_error_dic_10[k] = std_e
                mean_10[k] = (sum(num_list_p) - 0.1 * invalid) / (len(num_list_p) - invalid)
            std_error_dic_10x10[j] = std_error_dic_10
            mean_10x10[j] = mean_10
        std_error_dic_10x10x10[i] = std_error_dic_10x10
        mean_10x10x10[i] = mean_10x10
    return std_error_dic_10x10x10, mean_10x10x10


# key:seed
def json_p_to_std_error_dic(FILENAME):
    dic = json.load(open(FILENAME))
    std_error_dic_10x10 = {}
    mean_10x10 = {}

    for i in dic.keys():
        std_error_dic_10 = {}
        mean_10 = {}
        for j in dic[i].keys():
            invalid = 0
            num_list_seed = []
            for k in dic[i][j].keys():
                for l in dic[i][j][k].keys():
                    num_list_seed.append(dic[i][j][k][l])
                    if dic[i][j][k][l] == 0.1:
                        invalid += 1
            std_e = calculate_standard_error(num_list_seed, invalid)
            std_error_dic_10[j] = std_e
            mean_10[j] = (sum(num_list_seed) - 0.1 * invalid) / (len(num_list_seed) - invalid)
        std_error_dic_10x10[i] = std_error_dic_10
        mean_10x10[i] = mean_10

    return std_error_dic_10x10, mean_10x10


# key:ST
def json_seed_to_std_error_dic(FILENAME):
    dic = json.load(open(FILENAME))
    std_error_dic_10 = {}
    mean_dic_10 = {}

    for i in dic.keys():
        invalid = 0
        num_list_ST = []
        for j in dic[i].keys():
            for k in dic[i][j].keys():
                for l in dic[i][j][k].keys():
                    num_list_ST.append(dic[i][j][k][l])
                    if dic[i][j][k][l] == 0.1:
                        invalid += 1

        std_e = calculate_standard_error(num_list_ST, invalid)
        std_error_dic_10[i] = std_e
        if len(num_list_ST) - invalid == 0:
            continue
        else:
            mean_dic_10[i] = (sum(num_list_ST) - 0.1 * invalid) / (len(num_list_ST) - invalid)
    return std_error_dic_10, mean_dic_10


def json_dic_interval_to_length(FILENAME):
    dic = json.load(open(FILENAME))
    num_list = []
    for i in dic.keys():
        for j in dic[i].keys():
            for k in dic[i][j].keys():
                for l in dic[i][j][k].keys():
                    num_list.append(list(np.array(dic[i][j][k][l][1]) - np.array(dic[i][j][k][l][0])))

    mean = np.mean(num_list)
    std = np.std(num_list)
    return std, mean


def dic_to_mean_array(data, shape):
    total_sum = np.zeros(shape)
    count = 0
    # 遍历字典的所有层级并累加所有800维列表的值
    for st_key, st_value in data.items():
        for seed_key, seed_value in st_value.items():
            for p_key, p_value in seed_value.items():
                for count_key, count_value in p_value.items():
                    total_sum += np.array(count_value)

                    count += 1

    # 计算平均值
    average_sum = total_sum / count

    return average_sum


def list_to_dic(LIST):
    dic = {}
    count = 0
    key_list = list(range(0, len(LIST)))
    for k in key_list:
        dic[k] = LIST[count]
        count += 1
    return dic


def sort_label_interval(label, lower, upper):
    label = list_to_dic(label)
    lower = list_to_dic(lower)
    upper = list_to_dic(upper)

    sorted_keys = sorted(label, key=label.get)

    # 生成新字典 a_new 和 b_new
    label_sort = {key: label[key] for key in sorted_keys}
    lower_sort = {key: lower[key] for key in sorted_keys}
    upper_sort = {key: upper[key] for key in sorted_keys}

    label_sort = list(label_sort.values())
    lower_sort = list(lower_sort.values())
    upper_sort = list(upper_sort.values())

    return label_sort, lower_sort, upper_sort


def plot_label_and_interval2(a, lower, upper, interval_name, coverage):
    # 确定 x 坐标
    x = list(range(1, len(a) + 1))

    # 创建绘图
    plt.figure(figsize=(7, 5.6))

    # 绘制下界和上界的橙色散点
    plt.scatter(x, lower, color='orange', s=1, alpha=0.5,  label=interval_name + ' interval')
    plt.scatter(x, upper, color='orange', s=1, alpha=0.5)

    # 绘制真实值
    label_count1 = 0
    label_count2 = 0
    for i in range(len(a)):
        if lower[i] <= a[i] <= upper[i]:
            if label_count1 == 0:
                plt.scatter(x[i], a[i], color='green', s=3, label='edge_weight within interval')
                label_count1 = 1
            else:
                plt.scatter(x[i], a[i], color='green', s=3)
        else:
            if label_count2 == 0:
                plt.scatter(x[i], a[i], color='red', s=12, label='edge_weight out of interval')
                label_count2 = 1
            else:
                plt.scatter(x[i], a[i], color='red', s=12)

    # 设置标题和标签
    plt.title('Edge_weight and ' + interval_name + ' interval ' + '(coverage=' + str(coverage) + ')', fontsize=12)
    plt.xlabel('index')
    plt.ylabel('weight')

    # 添加图例
    plt.legend()

    # 显示图形
    plt.show()

    return 0



NAME = '\V4_comment_delta\Interval_NO_OPT2\V5_3_1_100'
file_cqr_interval = 'D:\OneDrive - City University of Hong Kong - Student\PyCharm202122\shortest_path202403\GNN\Data' + NAME + '\CQR_interval.json'
file_qr_interval = 'D:\OneDrive - City University of Hong Kong - Student\PyCharm202122\shortest_path202403\GNN\Data' + NAME + '\QR_interval.json'
file_bl_interval = 'D:\OneDrive - City University of Hong Kong - Student\PyCharm202122\shortest_path202403\GNN\Data' + NAME + '\BL_interval.json'
file_cqr_new_interval = 'D:\OneDrive - City University of Hong Kong - Student\PyCharm202122\shortest_path202403\GNN\Data' + NAME + '\CQR_NEW_interval.json'

file_c_label = 'D:\OneDrive - City University of Hong Kong - Student\PyCharm202122\shortest_path202403\GNN\Data' + NAME + '\C_LABEL.json'

file_cov_cqr = 'D:\OneDrive - City University of Hong Kong - Student\PyCharm202122\shortest_path202403\GNN\Data' + NAME + '\COV_CQR.json'
file_cov_cqr_new = 'D:\OneDrive - City University of Hong Kong - Student\PyCharm202122\shortest_path202403\GNN\Data' + NAME + '\COV_CQR_NEW.json'
file_cov_qr = 'D:\OneDrive - City University of Hong Kong - Student\PyCharm202122\shortest_path202403\GNN\Data' + NAME + '\COV_QR.json'
file_cov_bl = 'D:\OneDrive - City University of Hong Kong - Student\PyCharm202122\shortest_path202403\GNN\Data' + NAME + '\COV_BL.json'

########################## plot interval+label #####################################
cqr_new_interval = json.load(open(file_cqr_new_interval))
cqr_interval = json.load(open(file_cqr_interval))
qr_interval = json.load(open(file_qr_interval))
c_label_dic = json.load(open(file_c_label))

cqr_new_interval_array = dic_to_mean_array(cqr_new_interval, (2, 860))
cqr_interval_array = dic_to_mean_array(cqr_interval, (2, 860))
qr_interval_array = dic_to_mean_array(qr_interval, (2, 860))
c_label_array = dic_to_mean_array(c_label_dic, (1, 860))
#
# cqr_new_lower = cqr_new_interval_array[0].tolist()
# cqr_new_upper = cqr_new_interval_array[1].tolist()
# cqr_lower = cqr_interval_array[0].tolist()
# cqr_upper = cqr_interval_array[1].tolist()
# qr_lower = qr_interval_array[0].tolist()
# qr_upper = qr_interval_array[1].tolist()
# c_label = c_label_array.tolist()
#
# 0.9349
label_s, qr_lower_s, qr_upper_s = sort_label_interval(c_label_dic['(31,512)']['_seed_13877']['p_0.8']['shuffle_4'],
                                                      qr_interval['(31,512)']['_seed_13877']['p_0.8']['shuffle_4'][0],
                                                      qr_interval['(31,512)']['_seed_13877']['p_0.8']['shuffle_4'][1])

plot_label_and_interval2(label_s, qr_lower_s, qr_upper_s, 'qr', 0.9279)

# 0.9477
label_s, cqr_lower_s, cqr_upper_s = sort_label_interval(c_label_dic['(31,512)']['_seed_13877']['p_0.8']['shuffle_4'],
                                                        cqr_interval['(31,512)']['_seed_13877']['p_0.8']['shuffle_4'][
                                                            0],
                                                        cqr_interval['(31,512)']['_seed_13877']['p_0.8']['shuffle_4'][
                                                            1])

plot_label_and_interval2(label_s, cqr_lower_s, cqr_upper_s, 'cqr', 0.9512)

#0.9500
label_s, cqr_new_lower_s, cqr_new_upper_s = sort_label_interval(
    c_label_dic['(31,512)']['_seed_13877']['p_0.8']['shuffle_4'],
    cqr_new_interval['(31,512)']['_seed_13877']['p_0.8']['shuffle_4'][0],
    cqr_new_interval['(31,512)']['_seed_13877']['p_0.8']['shuffle_4'][1])

plot_label_and_interval2(label_s, cqr_new_lower_s, cqr_new_upper_s, 'cqr_erc', 0.95)

#
# x = np.arange(len(cqr_new_lower_s))
# parameter = np.polyfit(x, cqr_new_lower_s, 4)
#
# y2 = parameter[0] * x ** 4 + parameter[1] * x ** 3 + parameter[2] * x ** 2 + parameter[3] * x + parameter[4]
# plt.plot(x, y2, color='orange', label = 'lower_bound')
# plt.show()
##################################### mean std #############################################################
cov_qr = json.load(open(file_cov_qr))
cov_cqr = json.load(open(file_cov_cqr))
cov_cqr_new = json.load(open(file_cov_cqr_new))

